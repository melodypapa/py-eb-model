"""
Crc Excel Reporter Tests - Tests for CrcXdmXlsWriter.
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.mem_stack.crc_xdm import CrcXdmXlsWriter
from eb_model.models.mem_stack.crc_xdm import Crc, CrcConfig
from eb_model.models.core.eb_doc import EBModel


class TestCrcXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file."""
        writer = CrcXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            doc.getCrc()

            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_crc_config_with_data(self):
        """Test writing Crc configurations to Excel."""
        writer = CrcXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            crc = doc.getCrc()

            config1 = CrcConfig(crc, "Config1")
            config1.setCrcId(1)
            config1.setCrcCRCType("CRC8")

            config2 = CrcConfig(crc, "Config2")
            config2.setCrcId(2)
            config2.setCrcCRCType("CRC16")

            crc.addCrcConfig(config1)
            crc.addCrcConfig(config2)

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["CrcConfig"]

            # Check title row
            assert sheet.cell(1, 1).value == "Name"
            assert sheet.cell(1, 2).value == "CrcId"
            assert sheet.cell(1, 3).value == "CRCType"

            # Check data rows
            assert sheet.cell(2, 1).value == "Config1"
            assert sheet.cell(2, 2).value == 1
            assert sheet.cell(2, 3).value == "CRC8"

            assert sheet.cell(3, 1).value == "Config2"
            assert sheet.cell(3, 2).value == 2
            assert sheet.cell(3, 3).value == "CRC16"

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
