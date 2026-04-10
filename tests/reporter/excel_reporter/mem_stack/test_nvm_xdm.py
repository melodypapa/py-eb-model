"""
NvM Excel Reporter Tests.

Implements:
    - TC_UNIT_REPORTER_00017: NvM Excel Reporter - File Creation

Implements: SWR_REPORTER_00002
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.mem_stack.nvm_xdm import NvMXdmXlsWriter
from eb_model.models.core.eb_doc import EBModel
from eb_model.models.mem_stack.nvm_xdm import NvMCommon


class TestNvMXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file.

        Implements: TC_UNIT_REPORTER_00017
        """
        writer = NvMXdmXlsWriter()
        doc = EBModel.getInstance()

        # Create minimal NvM fixture
        nvm = doc.getNvM()
        nvm_common = NvMCommon(nvm, "NvMCommon")
        nvm_common.setNvMCompiledConfigId(0).setNvMDatasetSelectionBits(2)
        nvm.setNvMCommon(nvm_common)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            writer.write(filename, doc, options={})

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            assert len(wb.sheetnames) > 0
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
