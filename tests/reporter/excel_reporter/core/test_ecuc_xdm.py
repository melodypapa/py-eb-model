"""
EcuC Excel Reporter Tests.

Implements:
    - TC_UNIT_REPORTER_00030: EcuC Excel Reporter - File Creation

Implements: SWR_REPORTER_00002
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.core.ecuc_xdm import EcucXdmXlsWriter
from eb_model.models.core.eb_doc import EBModel
from eb_model.models.core.ecuc_xdm import EcucPartitionCollection, EcucPartition


class TestEcucXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file.

        Implements: TC_UNIT_REPORTER_00030
        """
        writer = EcucXdmXlsWriter()
        doc = EBModel.getInstance()

        # Create minimal EcuC fixture with partition collection
        ecuc = doc.getEcuC()
        partition_collection = EcucPartitionCollection(ecuc, "EcucPartitionCollection")
        partition = EcucPartition(partition_collection, "Partition_0")
        partition.setEcucPartitionId(0).setEcucPartitionCanBeRestarted(True)
        partition_collection.addEcucPartition(partition)
        ecuc.setEcucPartitionCollection(partition_collection)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            assert len(wb.sheetnames) > 0
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
