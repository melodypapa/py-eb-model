"""
PduR Excel Reporter Tests - Tests for PduRXdmXlsWriter.
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.com_stack.pdur_xdm import PduRXdmXlsWriter
from eb_model.models.com_stack.pdur_xdm import PduR, PduRRoutingTableEntry
from eb_model.models.core.eb_doc import EBModel
from eb_model.models.core.abstract import EcucRefType


class TestPduRXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file."""
        writer = PduRXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            doc.getPduR()

            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_pdu_r_routing_table_with_data(self):
        """Test writing PduR routing table to Excel."""
        writer = PduRXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            pdur = doc.getPduR()

            entry1 = PduRRoutingTableEntry(pdur, "RoutingEntry1")
            entry1.setPduRRoutingTableEntryID(1)
            entry1.setPduRRoutingPduSID(100)
            entry1.setPduRDestPduRef(EcucRefType("/Can/DestPdu"))
            entry1.setPduRSrcPduRef(EcucRefType("/Can/SrcPdu"))

            entry2 = PduRRoutingTableEntry(pdur, "RoutingEntry2")
            entry2.setPduRRoutingTableEntryID(2)
            entry2.setPduRRoutingPduSID(200)

            pdur.addPduRRoutingTableEntry(entry1)
            pdur.addPduRRoutingTableEntry(entry2)

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["PduRRoutingTable"]

            # Check title row
            assert sheet.cell(1, 1).value == "Name"
            assert sheet.cell(1, 2).value == "EntryID"
            assert sheet.cell(1, 3).value == "PduSID"
            assert sheet.cell(1, 4).value == "SrcPduRef"
            assert sheet.cell(1, 5).value == "DestPduRef"

            # Check data rows
            assert sheet.cell(2, 1).value == "RoutingEntry1"
            assert sheet.cell(2, 2).value == 1
            assert sheet.cell(2, 3).value == 100
            assert sheet.cell(2, 4).value == "/Can/SrcPdu"
            assert sheet.cell(2, 5).value == "/Can/DestPdu"

            assert sheet.cell(3, 1).value == "RoutingEntry2"
            assert sheet.cell(3, 2).value == 2
            assert sheet.cell(3, 3).value == 200

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
