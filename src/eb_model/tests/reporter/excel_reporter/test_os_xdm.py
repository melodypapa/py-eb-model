"""
Os Excel Reporter Tests.

Implements:
    - TC_UNIT_REPORTER_00016: Os Excel Reporter - File Creation

Implements: SWR_REPORTER_00002
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.os_xdm import OsXdmXlsWriter
from eb_model.models.core.eb_doc import EBModel


class TestOsXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file.

        Implements: TC_UNIT_REPORTER_00016
        """
        writer = OsXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            writer.write(filename, doc, options={"skip_os_task": True})

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            # OS reporter creates multiple sheets
            assert len(wb.sheetnames) > 0
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_creates_expected_sheets(self):
        """Test that write() creates expected sheets.

        Implements: TC_UNIT_REPORTER_00016
        """
        writer = OsXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            writer.write(filename, doc, options={"skip_os_task": True})

            wb = load_workbook(filename)

            # Check for expected sheets
            expected_sheets = ["OsApplications", "OsIsr"]
            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames, f"Sheet {sheet_name} not found"

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)