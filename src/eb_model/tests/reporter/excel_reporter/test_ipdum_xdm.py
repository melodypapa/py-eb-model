"""
IpduM Excel Reporter Tests - Tests for IpduMXdmXlsWriter.
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.ipdum_xdm import IpduMXdmXlsWriter
from eb_model.models.ipdum_xdm import IpduM, IpduMDynPdu
from eb_model.models.eb_doc import EBModel


class TestIpduMXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file."""
        writer = IpduMXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            doc.getIpduM()

            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_ipdu_m_dyn_pdu_with_data(self):
        """Test writing IpduM dynamic PDU to Excel."""
        writer = IpduMXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            ipdum = doc.getIpduM()

            pdu1 = IpduMDynPdu(ipdum, "DynPdu1")
            pdu1.setIpduMDynPduId(1)
            pdu1.setIpduMDynPduLength(8)

            pdu2 = IpduMDynPdu(ipdum, "DynPdu2")
            pdu2.setIpduMDynPduId(2)
            pdu2.setIpduMDynPduLength(16)

            ipdum.addIpduMDynPdu(pdu1)
            ipdum.addIpduMDynPdu(pdu2)

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["IpduMDynPdu"]

            # Check title row
            assert sheet.cell(1, 1).value == "Name"
            assert sheet.cell(1, 2).value == "DynPduId"
            assert sheet.cell(1, 3).value == "DynPduLength"

            # Check data rows
            assert sheet.cell(2, 1).value == "DynPdu1"
            assert sheet.cell(2, 2).value == 1
            assert sheet.cell(2, 3).value == 8

            assert sheet.cell(3, 1).value == "DynPdu2"
            assert sheet.cell(3, 2).value == 2
            assert sheet.cell(3, 3).value == 16

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
