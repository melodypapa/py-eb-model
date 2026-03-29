"""
EcuC Excel Reporter Tests.

Implements:
    - TC_UNIT_REPORTER_00030: EcuC Excel Reporter - File Creation

Implements: SWR_REPORTER_00002
"""
import os
import tempfile
from openpyxl import load_workbook
import pytest
from eb_model.reporter.excel_reporter.core.ecuc_xdm import EcucXdmXlsWriter
from eb_model.models.core.eb_doc import EBModel


class TestEcucXdmXlsWriter:

    @pytest.mark.skip(reason="EcuC reporter requires full EcuC partition collection setup")
    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file.

        Implements: TC_UNIT_REPORTER_00030
        """
        writer = EcucXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            assert len(wb.sheetnames) > 0
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)