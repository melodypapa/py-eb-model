"""
ComM Excel Reporter Tests - Tests for ComMXdmXlsWriter.
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.com_stack.comm_xdm import ComMXdmXlsWriter
from eb_model.models.com_stack.comm_xdm import ComM, ComMChannel
from eb_model.models.core.eb_doc import EBModel


class TestComMXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file."""
        writer = ComMXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            doc.getComM()

            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_com_m_channel_with_data(self):
        """Test writing ComM channels to Excel."""
        writer = ComMXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            comm = doc.getComM()

            channel1 = ComMChannel(comm, "Channel1")
            channel1.setComMChannelName("TestChannel1")
            channel1.setComMChannelId(1)

            channel2 = ComMChannel(comm, "Channel2")
            channel2.setComMChannelName("TestChannel2")
            channel2.setComMChannelId(2)

            comm.addComMChannel(channel1)
            comm.addComMChannel(channel2)

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["ComMChannel"]

            # Check title row
            assert sheet.cell(1, 1).value == "Name"
            assert sheet.cell(1, 2).value == "ChannelName"
            assert sheet.cell(1, 3).value == "ChannelId"

            # Check data rows
            assert sheet.cell(2, 1).value == "Channel1"
            assert sheet.cell(2, 2).value == "TestChannel1"
            assert sheet.cell(2, 3).value == 1

            assert sheet.cell(3, 1).value == "Channel2"
            assert sheet.cell(3, 2).value == "TestChannel2"
            assert sheet.cell(3, 3).value == 2

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
