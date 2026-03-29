"""
Nm Excel Reporter Tests - Tests for NmXdmXlsWriter.
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.nm_xdm import NmXdmXlsWriter
from eb_model.models.com_stack.nm_xdm import Nm, NmChannel
from eb_model.models.core.eb_doc import EBModel
from eb_model.models.core.abstract import EcucRefType


class TestNmXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file."""
        writer = NmXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            doc.getNm()

            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_nm_channel_with_data(self):
        """Test writing Nm channels to Excel."""
        writer = NmXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            nm = doc.getNm()

            channel1 = NmChannel(nm, "Channel1")
            channel1.setNmChannelId(1)
            channel1.setNmBusType("CAN")
            channel1.setNmMsgCycleTime(100.0)
            channel1.setNmTimeoutTime(200.0)
            channel1.setNmNetworkHandle(1)
            channel1.setNmComMNetworkHandleRef(EcucRefType("/ComM/Channel1"))
            channel1.setNmNodeEnabled(True)

            channel2 = NmChannel(nm, "Channel2")
            channel2.setNmChannelId(2)
            channel2.setNmBusType("ETH")

            nm.addNmChannel(channel1)
            nm.addNmChannel(channel2)

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["NmChannel"]

            # Check title row
            assert sheet.cell(1, 1).value == "Name"
            assert sheet.cell(1, 2).value == "ChannelId"
            assert sheet.cell(1, 3).value == "BusType"
            assert sheet.cell(1, 4).value == "MsgCycleTime"
            assert sheet.cell(1, 5).value == "TimeoutTime"
            assert sheet.cell(1, 6).value == "NetworkHandle"
            assert sheet.cell(1, 7).value == "ComMNetworkHandleRef"
            assert sheet.cell(1, 8).value == "NodeEnabled"

            # Check first data row
            assert sheet.cell(2, 1).value == "Channel1"
            assert sheet.cell(2, 2).value == 1
            assert sheet.cell(2, 3).value == "CAN"
            assert sheet.cell(2, 4).value == 100.0
            assert sheet.cell(2, 5).value == 200.0
            assert sheet.cell(2, 6).value == 1
            assert sheet.cell(2, 7).value == "/ComM/Channel1"
            assert sheet.cell(2, 8).value == "Y"

            # Check second data row
            assert sheet.cell(3, 1).value == "Channel2"
            assert sheet.cell(3, 2).value == 2
            assert sheet.cell(3, 3).value == "ETH"

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
