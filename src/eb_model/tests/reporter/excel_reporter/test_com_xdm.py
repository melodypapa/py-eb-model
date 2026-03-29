"""
Com Excel Reporter Tests - Tests for ComXdmXlsWriter.
"""
import os
import tempfile
from openpyxl import load_workbook
from eb_model.reporter.excel_reporter.com_xdm import ComXdmXlsWriter
from eb_model.models.com_stack.com_xdm import Com, ComGeneral
from eb_model.models.core.eb_doc import EBModel


class TestComXdmXlsWriter:

    def test_write_creates_excel_file(self):
        """Test that write() creates a valid Excel file."""
        writer = ComXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            com = doc.getCom()
            general = ComGeneral(com, "ComGeneral")
            general.setComEnableUserSupport(True)
            com.setComGeneral(general)

            writer.write(filename, doc)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            wb = load_workbook(filename)
            assert "ComGeneral" in wb.sheetnames
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_com_general_with_data(self):
        """Test writing ComGeneral configuration to Excel."""
        writer = ComXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            com = doc.getCom()
            general = ComGeneral(com, "ComGeneral")
            general.setComEnableUserSupport(True)
            general.setComUserInitSignal(False)
            general.setComUserStatusSupport(True)
            general.setComUserTxConfirmation(False)
            general.setComUserRxIndication(True)
            com.setComGeneral(general)

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["ComGeneral"]

            # Check title row
            assert sheet.cell(1, 1).value == "Name"
            assert sheet.cell(1, 2).value == "EnableUserSupport"
            assert sheet.cell(1, 3).value == "UserInitSignal"
            assert sheet.cell(1, 4).value == "UserStatusSupport"
            assert sheet.cell(1, 5).value == "UserTxConfirmation"
            assert sheet.cell(1, 6).value == "UserRxIndication"

            # Check data row (booleans are written as "Y" for True, None for False)
            assert sheet.cell(2, 1).value == "ComGeneral"
            assert sheet.cell(2, 2).value == "Y"
            assert sheet.cell(2, 3).value is None
            assert sheet.cell(2, 4).value == "Y"
            assert sheet.cell(2, 5).value is None
            assert sheet.cell(2, 6).value == "Y"

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_write_com_general_empty(self):
        """Test writing when ComGeneral is None."""
        writer = ComXdmXlsWriter()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            doc = EBModel.getInstance()
            # Don't set ComGeneral, keep it None

            writer.write(filename, doc)

            wb = load_workbook(filename)
            sheet = wb["ComGeneral"]

            # Title row should exist
            assert sheet.cell(1, 1).value == "Name"

            # Even when ComGeneral is None, the reporter writes the row with None values
            # This is expected behavior from the abstract ExcelReporter
            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)
