"""
Excel Reporter Abstract Base Class Tests.

Implements:
    - TC_UNIT_REPORTER_00001: Excel Reporter - Initialization
    - TC_UNIT_REPORTER_00002: Excel Reporter - Cell Formatting
    - TC_UNIT_REPORTER_00003: Excel Reporter - Auto-Width Formatting
    - TC_UNIT_REPORTER_00004: Excel Reporter - Boolean Formatting
    - TC_UNIT_REPORTER_00005: Excel Reporter - Workbook Save
    - TC_UNIT_REPORTER_00008: Reporter Layer - Error Handling

Implements: SWR_REPORTER_00001, SWR_REPORTER_00002, SWR_REPORTER_00003, SWR_REPORTER_00004, SWR_REPORTER_00009
"""
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pytest

from eb_model.reporter.excel_reporter.abstract import ExcelReporter


class TestExcelReporter:

    def test_initialization(self):
        """Test that ExcelReporter initializes correctly.

        Implements: TC_UNIT_REPORTER_00001
        """
        reporter = ExcelReporter()

        assert reporter.wb is not None
        assert reporter.logger is not None
        # Workbook has a default sheet
        assert len(reporter.wb.sheetnames) >= 1

    def test_write_cell(self):
        """Test writing cells with various values.

        Implements: TC_UNIT_REPORTER_00002
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write a normal value
        cell = reporter.write_cell(sheet, 1, 1, "Test Data")
        assert cell.value == "Test Data"

        # Write None value
        cell = reporter.write_cell(sheet, 2, 1, None)
        assert cell.value is None

        # Write empty string
        cell = reporter.write_cell(sheet, 3, 1, "")
        assert cell.value == ""

        # Write with alignment format
        cell = reporter.write_cell(sheet, 4, 1, "Centered", format={"alignment": Alignment(horizontal="center")})
        assert cell.alignment.horizontal == "center"

    def test_write_cell_with_number_format(self):
        """Test writing cells with number format.

        Implements: TC_UNIT_REPORTER_00002
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write with number format
        cell = reporter.write_cell(sheet, 1, 1, 123.456, format={"number_format": "0.00"})
        assert cell.value == 123.456
        assert cell.number_format == "0.00"

    def test_write_cell_center(self):
        """Test writing cells with center alignment.

        Implements: TC_UNIT_REPORTER_00002
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        reporter.write_cell_center(sheet, 1, 1, "Centered Text")
        assert sheet.cell(1, 1).value == "Centered Text"
        assert sheet.cell(1, 1).alignment.horizontal == "center"

    def test_write_title_row(self):
        """Test writing title row with center alignment.

        Implements: TC_UNIT_REPORTER_00002
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        title_row = ["Column A", "Column B", "Column C"]
        reporter.write_title_row(sheet, title_row)

        assert sheet.cell(1, 1).value == "Column A"
        assert sheet.cell(1, 2).value == "Column B"
        assert sheet.cell(1, 3).value == "Column C"

        assert sheet.cell(1, 1).alignment.horizontal == "center"
        assert sheet.cell(1, 2).alignment.horizontal == "center"
        assert sheet.cell(1, 3).alignment.horizontal == "center"

    def test_format_boolean(self):
        """Test boolean formatting.

        Implements: TC_UNIT_REPORTER_00004
        """
        reporter = ExcelReporter()

        assert reporter.format_boolean(True) == "Y"
        assert reporter.format_boolean(False) == ""

    def test_write_bool_cell(self):
        """Test writing boolean cells.

        Implements: TC_UNIT_REPORTER_00004
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write True
        reporter.write_bool_cell(sheet, 1, 1, True)
        assert sheet.cell(1, 1).value == "Y"
        assert sheet.cell(1, 1).alignment.horizontal == "center"

        # Write False
        reporter.write_bool_cell(sheet, 2, 1, False)
        assert sheet.cell(2, 1).value == ""
        assert sheet.cell(2, 1).alignment.horizontal == "center"

    def test_auto_width(self):
        """Test auto-width column formatting.

        Implements: TC_UNIT_REPORTER_00003
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write data of varying lengths
        reporter.write_cell(sheet, 1, 1, "Short")
        reporter.write_cell(sheet, 2, 1, "Medium Length")
        reporter.write_cell(sheet, 3, 1, "This is a very long string that should cause the column to expand")

        reporter.auto_width(sheet)

        # Column width should be greater than 0
        assert sheet.column_dimensions['A'].width > 0

    def test_auto_width_with_customized(self):
        """Test auto-width with customized column widths.

        Implements: TC_UNIT_REPORTER_00003
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write data
        reporter.write_cell(sheet, 1, 1, "Short")
        reporter.write_cell(sheet, 1, 2, "Medium Length")

        # Customize column A
        customized = {'A': 50}
        reporter.auto_width(sheet, customized=customized)

        # Column A should use customized width
        assert sheet.column_dimensions['A'].width == 50
        # Column B should use auto-width
        assert sheet.column_dimensions['B'].width > 0

    def test_auto_width_with_zero_customized(self):
        """Test auto-width with zero width customization.

        Implements: TC_UNIT_REPORTER_00003
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write data
        reporter.write_cell(sheet, 1, 1, "Long string here")

        # Customize column A with zero width
        # Note: The implementation doesn't set width to 0, it just leaves it unset
        # This test verifies the behavior
        customized = {'A': 0}
        reporter.auto_width(sheet, customized=customized)

        # When customized width is 0, it doesn't set it, so auto-width is used
        assert sheet.column_dimensions['A'].width > 0

    def test_auto_width_with_customized_positive_width(self):
        """Test auto-width with positive customized width.

        Implements: TC_UNIT_REPORTER_00003
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write data
        reporter.write_cell(sheet, 1, 1, "Long string here")

        # Customize column A with positive width
        customized = {'A': 50}
        reporter.auto_width(sheet, customized=customized)

        # Column A should use customized width
        assert sheet.column_dimensions['A'].width == 50

    def test_save_workbook(self):
        """Test saving workbook to file.

        Implements: TC_UNIT_REPORTER_00005
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write some data
        reporter.write_cell(sheet, 1, 1, "Test Content")
        reporter.write_cell(sheet, 2, 1, "More Content")

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            reporter.save(filename)

            assert os.path.exists(filename)
            assert os.path.getsize(filename) > 0

            # Load and verify
            wb = load_workbook(filename)
            loaded_sheet = wb.active

            assert loaded_sheet.cell(1, 1).value == "Test Content"
            assert loaded_sheet.cell(2, 1).value == "More Content"

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)

    def test_save_to_invalid_path(self):
        """Test saving to invalid path raises error.

        Implements: TC_UNIT_REPORTER_00008
        """
        reporter = ExcelReporter()

        # Try to save to a non-existent directory
        with pytest.raises(OSError):
            reporter.save("/nonexistent/directory/test.xlsx")

    def test_workbook_can_be_loaded_after_save(self):
        """Test that saved workbook can be loaded.

        Implements: TC_UNIT_REPORTER_00005
        """
        reporter = ExcelReporter()
        sheet = reporter.wb.active

        # Write various data types
        reporter.write_cell(sheet, 1, 1, "String")
        reporter.write_cell(sheet, 2, 1, 123)
        reporter.write_cell(sheet, 3, 1, True)
        reporter.write_cell(sheet, 4, 1, None)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            reporter.save(filename)

            # Load and verify all data
            wb = load_workbook(filename)
            loaded_sheet = wb.active

            assert loaded_sheet.cell(1, 1).value == "String"
            assert loaded_sheet.cell(2, 1).value == 123
            assert loaded_sheet.cell(3, 1).value is True
            assert loaded_sheet.cell(4, 1).value is None

            wb.close()
        finally:
            if os.path.exists(filename):
                os.remove(filename)